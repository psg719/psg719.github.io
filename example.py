from flask import Flask, redirect, url_for, render_template, request, session, flash
from datetime import timedelta
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



wb = load_workbook(filename = 'ingred.xlsx')
ws = wb['ingred']
drugs = []
for cell in ws['a']: #puts all values in x column in a list
    drugs.append(cell.value)
drugs.sort()

app = Flask(__name__)
app.secret_key = "hello"
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.sqlite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


class bdata(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    med1 = db.Column(db.String(120), unique=True, nullable=False)
    med1dir = db.Column(db.String(120), unique=True, nullable=False)

    def __init__(self,user,email,med1,med1dir):
        self.user = user
        self.email = email
        self.med1 = med1
        self.med1dir = med1dir

@app.route("/drop", methods=["POST","GET"])
def drop():
    return render_template("drop.html",drugs=drugs)

@app.route("/view")
def view():
    return render_template("view.html", values=bdata.query.all())

@app.route("/login", methods =["POST","GET"])
def login():
    return render_template("login.html", drugs=drugs)
    if request.method == "POST":
        session.permanent = True
        name = request.form["nm"]
        email = request.form["eml"]
        med1 = request.form["med1"]
        med1dir = request.form["med1dir"]
        #session["user"] = user
        usr = bdata(name,email,med1,med1dir)
        #eml = users(email, "")
        db.session.add(usr)
        db.session.commit()
        return render_template("view.html",values=bdata.query.all())
    else:
        return render_template("login.html")


if __name__ == "__main__":
    app.run(debug=True)
    db.create_all()


    '''    found_user = users.query.filter_by(name=user).first()
        if found_user:
            session["email"] = found_user.email
        else:'''





#for debug mode: <set FLASK_ENV=development> in CMD
